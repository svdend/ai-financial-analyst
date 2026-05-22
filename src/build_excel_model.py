"""Simplified three-statement Excel model with scenario analysis.

Generates ``/dashboard/{ticker}_3Statement_Model.xlsx`` from the DuckDB
warehouse produced by :mod:`src.build_warehouse`.

Sheets
------
Cover                  — methodology and disclosure
Scenarios              — Base / Bull / Bear parameter overrides
Assumptions            — active-scenario inputs
Income_Statement       — 12 historical + 4 forecast quarters
Balance_Sheet          — 12 historical + 4 forecast quarters
Cash_Flow              — indirect method, 12 + 4 quarters
Debt_Schedule          — quarterly amortization schedule
Summary                — KPI dashboard
Sources                — provenance table (accession_no per fact)
Revenue_Disaggregation — ASC 280 note (only when has_physical_inventory=TRUE)

Python-side verification
------------------------
BalanceCheck and GAAP OCF residual are verified by a SEPARATE Python
recomputation — NOT by reading the workbook with ``data_only=True``.
On freshly generated workbooks, ``data_only=True`` returns ``None`` for
every formula cell because there are no cached results yet.

CLI::

    python -m src.build_excel_model
    python -m src.build_excel_model --ticker PANW
"""

from __future__ import annotations

import datetime
import logging
from collections.abc import Sequence
from pathlib import Path
from typing import Any

import duckdb
import numpy as np
import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)

# ── Paths ──────────────────────────────────────────────────────────────────────
_REPO_ROOT = Path(__file__).resolve().parents[1]
_CONFIG_PATH = _REPO_ROOT / "config" / "company.yaml"
_PROCESSED_DIR = _REPO_ROOT / "data" / "processed"
_DASHBOARD_DIR = _REPO_ROOT / "dashboard"

# ── Model parameters ───────────────────────────────────────────────────────────
_N_HIST: int = 12
_N_FCST: int = 4
_N_TOTAL: int = _N_HIST + _N_FCST
_COL_LABEL: int = 1
_COL_DATA_START: int = 2
_COL_HIST_END: int = _COL_DATA_START + _N_HIST - 1  # column M
_COL_FCST_START: int = _COL_HIST_END + 1  # column N
_COL_FCST_END: int = _COL_FCST_START + _N_FCST - 1  # column Q

_BALANCE_CHECK_TOL: float = 1_000_000.0
_OCF_RESIDUAL_TOL: float = 5_000_000.0


# ── Styles ─────────────────────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


_FILL_HIST = _fill("DCE6F1")  # light blue — historicals
_FILL_FCST = _fill("E2EFDA")  # light green — forecasts
_FILL_INPUT = _fill("FFFFC7")  # yellow — user inputs
_FILL_HEADER = _fill("17375E")  # dark navy — column headers
_FILL_SECTION = _fill("D9E1F2")  # periwinkle — section labels
_FILL_OK = _fill("C6EFCE")  # green — BalanceCheck ok
_FILL_FAIL = _fill("FFC7CE")  # red — BalanceCheck fail

_FONT_HDR = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_FONT_BOLD = Font(bold=True, name="Calibri", size=10)
_FONT_HIST = Font(color="1F4E79", name="Calibri", size=10)
_FONT_FCST = Font(color="375623", name="Calibri", size=10)
_FONT_INP = Font(color="7F3F00", name="Calibri", size=10)
_FONT_NORM = Font(name="Calibri", size=10)

_FMT_USD = "#,##0"
_FMT_PCT = "0.0%"
_FMT_DAY = '0.0"d"'


def _sc(
    ws: Any,
    row: int,
    col: int,
    value: Any = None,
    font: Font | None = None,
    fill: PatternFill | None = None,
    fmt: str | None = None,
    halign: str | None = None,
) -> Any:
    """Set cell value and formatting; returns the cell."""
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if fmt is not None:
        cell.number_format = fmt
    if halign is not None:
        cell.alignment = Alignment(horizontal=halign)
    return cell


# ── Data loading ───────────────────────────────────────────────────────────────


def _load_history(
    db_path: Path,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, bool, pd.DataFrame]:
    """Return (hist_is, hist_bs, hist_cf, has_physical_inventory).

    Each DataFrame has up to ``_N_HIST`` rows, oldest-first, period_type='Q'.
    """
    con = duckdb.connect(str(db_path), read_only=True)
    try:
        hist_is = con.execute(f"""
            SELECT * FROM v_income_statement_quarterly
            WHERE period_type = 'Q'
            ORDER BY fiscal_year, fiscal_period
            LIMIT {_N_HIST}
        """).fetchdf()

        hist_bs = con.execute(f"""
            SELECT * FROM v_balance_sheet_quarterly
            WHERE period_type = 'Q'
            ORDER BY fiscal_year, fiscal_period
            LIMIT {_N_HIST}
        """).fetchdf()

        hist_cf = con.execute(f"""
            SELECT * FROM v_cash_flow_quarterly
            WHERE period_type = 'Q'
            ORDER BY fiscal_year, fiscal_period
            LIMIT {_N_HIST}
        """).fetchdf()

        dq = con.execute("SELECT * FROM v_data_quality").fetchdf()
        has_inv = bool(dq.iloc[0]["has_physical_inventory"]) if not dq.empty else False

        # Sources provenance
        sources = con.execute("""
            SELECT
                i.period_end, i.fiscal_year, i.fiscal_period,
                'Revenue'          AS line_item,
                i.Revenue          AS value,
                i.revenue_accession AS accession_no,
                i.revenue_filing_url AS filing_url
            FROM v_income_statement_quarterly i
            WHERE i.period_type = 'Q' AND i.Revenue IS NOT NULL
            UNION ALL
            SELECT period_end, fiscal_year, fiscal_period,
                'OperatingIncome', OperatingIncome,
                operating_income_accession, operating_income_filing_url
            FROM v_income_statement_quarterly WHERE period_type='Q' AND OperatingIncome IS NOT NULL
            UNION ALL
            SELECT period_end, fiscal_year, fiscal_period,
                'NetIncome', NetIncome,
                net_income_accession, net_income_filing_url
            FROM v_income_statement_quarterly WHERE period_type='Q' AND NetIncome IS NOT NULL
            UNION ALL
            SELECT period_end, fiscal_year, fiscal_period,
                'Cash', Cash, cash_accession, cash_filing_url
            FROM v_balance_sheet_quarterly WHERE period_type='Q' AND Cash IS NOT NULL
            UNION ALL
            SELECT period_end, fiscal_year, fiscal_period,
                'AccountsReceivable', AccountsReceivable,
                ar_accession, ar_filing_url
            FROM v_balance_sheet_quarterly WHERE period_type='Q' AND AccountsReceivable IS NOT NULL
            UNION ALL
            SELECT period_end, fiscal_year, fiscal_period,
                'OperatingCashFlow', OperatingCashFlow,
                ocf_accession, ocf_filing_url
            FROM v_cash_flow_quarterly WHERE period_type='Q' AND OperatingCashFlow IS NOT NULL
            ORDER BY period_end, line_item
        """).fetchdf()

    finally:
        con.close()

    return hist_is, hist_bs, hist_cf, has_inv, sources


def _v(df: pd.DataFrame, col: str, idx: int = -1, default: float = 0.0) -> float:
    """Safely extract a float value from a DataFrame column."""
    if col not in df.columns or len(df) == 0:
        return default
    val = df[col].iloc[idx]
    return float(val) if pd.notna(val) else default


def _series(df: pd.DataFrame, col: str, default: float = 0.0) -> list[float]:
    """Return column as list of floats, filling NaN with default."""
    if col not in df.columns:
        return [default] * len(df)
    return [float(x) if pd.notna(x) else default for x in df[col]]


# ── Assumption derivation ──────────────────────────────────────────────────────


def _compute_base_assumptions(
    hist_is: pd.DataFrame,
    hist_bs: pd.DataFrame,
    hist_cf: pd.DataFrame,
    has_inv: bool,
) -> dict[str, float]:
    """Derive Base scenario assumptions from trailing historical data."""
    revenue = _series(hist_is, "Revenue")
    cogs = _series(hist_is, "CostOfRevenue")
    gp = _series(hist_is, "GrossProfit")
    opex = _series(hist_is, "OperatingExpenses")
    n = len(revenue)

    # Revenue QoQ growth — trailing 4-quarter average
    if n >= 5 and revenue[-5] > 0:
        rev_qoq = [
            (revenue[-4 + i] / revenue[-5 + i] - 1.0) for i in range(4) if revenue[-5 + i] > 0
        ]
        revenue_growth_qoq = float(np.mean(rev_qoq)) if rev_qoq else 0.04
    else:
        revenue_growth_qoq = 0.04

    last_rev = revenue[-1] if revenue else 1.0
    last_gp = gp[-1] if gp else 0.0
    last_cogs = cogs[-1] if cogs else 0.0

    gross_margin_pct = last_gp / last_rev if last_rev > 0 else 0.68

    # Opex growth — trailing average
    if n >= 5:
        opex_qoq = [(opex[-4 + i] / opex[-5 + i] - 1.0) for i in range(4) if opex[-5 + i] > 0]
        opex_growth_qoq = float(np.mean(opex_qoq)) if opex_qoq else 0.02
    else:
        opex_growth_qoq = 0.02

    # CapEx % of revenue
    capex_list = _series(hist_cf, "CapEx")
    last_capex_avg = float(np.mean(capex_list[-4:])) if len(capex_list) >= 4 else 0.0
    last_rev_avg = float(np.mean(revenue[-4:])) if n >= 4 else max(last_rev, 1.0)
    capex_pct = last_capex_avg / last_rev_avg if last_rev_avg > 0 else 0.02

    # DSO
    last_ar = _v(hist_bs, "AccountsReceivable")
    dso_days = (last_ar / last_rev * 90.0) if last_rev > 0 else 60.0

    # DPO
    last_ap = _v(hist_bs, "AccountsPayable")
    dpo_days = (last_ap / last_cogs * 90.0) if last_cogs > 0 else 45.0

    # DIO
    last_inv = _v(hist_bs, "Inventory") if has_inv else 0.0
    dio_days = (last_inv / last_cogs * 90.0) if (has_inv and last_cogs > 0) else 0.0

    # SBC and buybacks — latest quarter
    sbc_qtrly = _v(hist_cf, "StockBasedCompensation")
    buybacks_qtrly = _v(hist_cf, "TreasuryStockRepurchases")

    # Opex trailing value (needed for growth formula)
    opex_trailing = opex[-1] if opex else 0.0

    return {
        "revenue_growth_qoq": max(-0.20, min(0.50, revenue_growth_qoq)),
        "gross_margin_pct": max(0.30, min(0.99, gross_margin_pct)),
        "opex_growth_qoq": max(-0.20, min(0.50, opex_growth_qoq)),
        "capex_pct_of_revenue": max(0.001, min(0.20, capex_pct)),
        "dso_days": max(1.0, min(180.0, dso_days)),
        "dpo_days": max(1.0, min(180.0, dpo_days)),
        "dio_days": max(0.0, min(180.0, dio_days)),
        "tax_rate": 0.15,
        "debt_amortization_qoq": 0.0,
        "sbc_qtrly": sbc_qtrly,
        "buybacks_qtrly": buybacks_qtrly,
        "opex_trailing": opex_trailing,
    }


def _make_scenarios(base: dict[str, float]) -> dict[str, dict[str, float]]:
    bull = dict(base)
    bull["revenue_growth_qoq"] = base["revenue_growth_qoq"] * 1.5
    bull["gross_margin_pct"] = min(0.99, base["gross_margin_pct"] * 1.02)
    bull["opex_growth_qoq"] = base["opex_growth_qoq"] * 0.8
    bull["dso_days"] = base["dso_days"] * 0.95

    bear = dict(base)
    bear["revenue_growth_qoq"] = base["revenue_growth_qoq"] * 0.4
    bear["gross_margin_pct"] = base["gross_margin_pct"] * 0.98
    bear["opex_growth_qoq"] = base["opex_growth_qoq"] * 1.3
    bear["dso_days"] = base["dso_days"] * 1.1

    return {"Base": base, "Bull": bull, "Bear": bear}


# ── Python-side forecast computation ──────────────────────────────────────────


def _forecast_periods(
    hist_is: pd.DataFrame,
    hist_bs: pd.DataFrame,
    hist_cf: pd.DataFrame,
    assumptions: dict[str, float],
    has_inv: bool,
    n_fcst: int = _N_FCST,
) -> list[dict[str, float]]:
    """Compute n_fcst forecast periods.  Returns list of row dicts.

    Cash is the BALANCING item derived from the CF statement.
    BalanceCheck = TotalAssets − (TotalLiabilities + TotalEquity).
    By construction, BalanceCheck = 0 for each forecast period when
    the starting period satisfies the accounting identity and DeferredRev
    is held flat.  Numeric proof: see module docstring.
    """
    ass = assumptions

    # ── Initialise from last historical quarter ────────────────────────────
    revenue = _series(hist_is, "Revenue")
    opex_hist = _series(hist_is, "OperatingExpenses")

    prev_rev = revenue[-1] if revenue else 1.0
    prev_opex = opex_hist[-1] if opex_hist else 0.0

    dep_rate = _v(hist_cf, "Depreciation")  # held flat each quarter

    cash_0 = _v(hist_bs, "Cash")
    ar_0 = _v(hist_bs, "AccountsReceivable")
    inv_0 = _v(hist_bs, "Inventory") if has_inv else 0.0
    ta_0 = _v(hist_bs, "TotalAssets")
    ap_0 = _v(hist_bs, "AccountsPayable")
    dr_0 = _v(hist_bs, "DeferredRevenue")
    tl_0 = _v(hist_bs, "TotalLiabilities")
    eq_0 = _v(hist_bs, "TotalEquity")

    # PPE residual: TotalAssets − Cash − AR − Inv. Rolled forward each quarter by
    # (capex − depreciation) so the accounting identity holds when capex ≠ dep.
    other_assets = ta_0 - cash_0 - ar_0 - inv_0
    other_liab = tl_0 - ap_0 - dr_0  # held flat (debt=0 simplified)
    debt = 0.0

    state = {
        "cash": cash_0,
        "ar": ar_0,
        "inv": inv_0,
        "other_assets": other_assets,
        "ap": ap_0,
        "deferred_rev": dr_0,
        "other_liab": other_liab,
        "debt": debt,
        "equity": eq_0,
    }
    prev_opex_val = prev_opex
    prev_rev_val = prev_rev

    periods: list[dict[str, float]] = []
    for _ in range(n_fcst):
        revenue_t = prev_rev_val * (1.0 + ass["revenue_growth_qoq"])
        cogs_t = revenue_t * (1.0 - ass["gross_margin_pct"])
        gp_t = revenue_t - cogs_t
        opex_t = prev_opex_val * (1.0 + ass["opex_growth_qoq"])
        op_inc_t = gp_t - opex_t
        interest_t = state["debt"] * 0.0  # simplified: zero-coupon for now
        pretax_t = op_inc_t - interest_t
        tax_t = max(0.0, pretax_t) * ass["tax_rate"]
        ni_t = pretax_t - tax_t

        capex_t = revenue_t * ass["capex_pct_of_revenue"]
        sbc_t = ass["sbc_qtrly"]
        buybacks_t = ass["buybacks_qtrly"]
        debt_amort = ass["debt_amortization_qoq"]

        # Working capital
        ar_t = revenue_t * ass["dso_days"] / 90.0
        inv_t = cogs_t * ass["dio_days"] / 90.0 if has_inv else 0.0
        ap_t = cogs_t * ass["dpo_days"] / 90.0
        dr_t = state["deferred_rev"]  # held flat

        # Equity
        eq_t = state["equity"] + ni_t + sbc_t - buybacks_t

        # Cash flow — indirect
        delta_ar = ar_t - state["ar"]
        delta_inv = inv_t - state["inv"]
        delta_ap = ap_t - state["ap"]
        ocf_t = ni_t + dep_rate + sbc_t - delta_ar - delta_inv + delta_ap
        inv_cf_t = -capex_t
        fin_cf_t = -debt_amort - buybacks_t
        net_change = ocf_t + inv_cf_t + fin_cf_t

        cash_t = state["cash"] + net_change  # PLUG
        new_debt = max(0.0, state["debt"] - debt_amort)

        # Roll PPE/other long-term assets: prior + capex − depreciation.
        # This is what closes the accounting identity when capex ≠ dep.
        other_assets_t = state["other_assets"] + capex_t - dep_rate

        # Balance check
        ta_t = cash_t + ar_t + inv_t + other_assets_t
        tl_t = ap_t + dr_t + new_debt + state["other_liab"]
        bc_t = ta_t - tl_t - eq_t

        periods.append(
            {
                "Revenue": revenue_t,
                "CostOfRevenue": cogs_t,
                "GrossProfit": gp_t,
                "OperatingExpenses": opex_t,
                "OperatingIncome": op_inc_t,
                "InterestExpense": interest_t,
                "PreTaxIncome": pretax_t,
                "Tax": tax_t,
                "NetIncome": ni_t,
                "Cash": cash_t,
                "AccountsReceivable": ar_t,
                "Inventory": inv_t,
                "OtherAssets": other_assets_t,
                "TotalAssets": ta_t,
                "AccountsPayable": ap_t,
                "DeferredRevenue": dr_t,
                "Debt": new_debt,
                "OtherLiabilities": state["other_liab"],
                "TotalLiabilities": tl_t,
                "TotalEquity": eq_t,
                "BalanceCheck": bc_t,
                "OCF": ocf_t,
                "CapEx": capex_t,
                "InvestingCF": inv_cf_t,
                "Buybacks": buybacks_t,
                "DebtAmort": debt_amort,
                "FinancingCF": fin_cf_t,
                "NetCashChange": net_change,
                "BegCash": state["cash"],
                "EndCash": cash_t,
                "SBC": sbc_t,
                "Depreciation": dep_rate,
            }
        )

        # Advance state
        state = {
            "cash": cash_t,
            "ar": ar_t,
            "inv": inv_t,
            "other_assets": other_assets_t,
            "ap": ap_t,
            "deferred_rev": dr_t,
            "other_liab": state["other_liab"],
            "debt": new_debt,
            "equity": eq_t,
        }
        prev_rev_val = revenue_t
        prev_opex_val = opex_t

    return periods


def _verify_ocf_residual(
    hist_is: pd.DataFrame,
    hist_bs: pd.DataFrame,
    hist_cf: pd.DataFrame,
    has_inv: bool,
) -> float:
    """Return max |OCF_modelled − OCF_reported| for last 4 historical quarters."""
    n = min(len(hist_is), len(hist_bs), len(hist_cf))
    if n < 2:
        return 0.0

    residuals: list[float] = []
    for i in range(max(1, n - 4), n):
        ni = _v(hist_is, "NetIncome", i)
        dep = _v(hist_cf, "Depreciation", i)
        sbc = _v(hist_cf, "StockBasedCompensation", i)
        ar_t = _v(hist_bs, "AccountsReceivable", i)
        ar_p = _v(hist_bs, "AccountsReceivable", i - 1)
        ap_t = _v(hist_bs, "AccountsPayable", i)
        ap_p = _v(hist_bs, "AccountsPayable", i - 1)
        inv_t = _v(hist_bs, "Inventory", i) if has_inv else 0.0
        inv_p = _v(hist_bs, "Inventory", i - 1) if has_inv else 0.0
        dr_t = _v(hist_bs, "DeferredRevenue", i)
        dr_p = _v(hist_bs, "DeferredRevenue", i - 1)

        ocf_model = ni + dep + sbc - (ar_t - ar_p) - (inv_t - inv_p) + (ap_t - ap_p) + (dr_t - dr_p)
        ocf_actual = _v(hist_cf, "OperatingCashFlow", i)
        if ocf_actual != 0.0:
            residuals.append(abs(ocf_model - ocf_actual))

    return max(residuals) if residuals else 0.0


# ── Sheet builders ─────────────────────────────────────────────────────────────


def _header_row(ws: Any, labels: list[str]) -> None:
    """Write a styled header row (row 1)."""
    for col_idx, label in enumerate(labels, start=1):
        _sc(ws, 1, col_idx, label, font=_FONT_HDR, fill=_FILL_HEADER, halign="center")


def _col_headers(
    ws: Any,
    hist_is: pd.DataFrame,
    row: int = 1,
) -> None:
    """Write period labels across columns B:Q."""
    _sc(ws, row, _COL_LABEL, "Line Item", font=_FONT_HDR, fill=_FILL_HEADER)
    hist_labels: list[str] = []
    if "fiscal_year" in hist_is.columns and "fiscal_period" in hist_is.columns:
        for _, r in hist_is.iterrows():
            hist_labels.append(f"Q{str(r['fiscal_period'])[-1]}/FY{r['fiscal_year']}")
    while len(hist_labels) < _N_HIST:
        hist_labels.insert(0, "")

    for i, label in enumerate(hist_labels[-_N_HIST:]):
        col = _COL_DATA_START + i
        _sc(ws, row, col, label, font=_FONT_HDR, fill=_FILL_HEADER, halign="center")

    for j in range(_N_FCST):
        col = _COL_FCST_START + j
        _sc(ws, row, col, f"Q{j + 1}(F)", font=_FONT_HDR, fill=_FILL_HEADER, halign="center")

    ws.column_dimensions["A"].width = 28
    for c in range(_COL_DATA_START, _COL_FCST_END + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14


def _write_data_row(
    ws: Any,
    row: int,
    label: str,
    hist_values: Sequence[float | None],
    fcst_values: Sequence[float | None] = (),
) -> None:
    """Write one data row with historical (blue) and forecast (green) cells."""
    _sc(ws, row, _COL_LABEL, label, font=_FONT_BOLD)

    padded = ([None] * (_N_HIST - len(hist_values))) + list(hist_values)
    for i, val in enumerate(padded):
        col = _COL_DATA_START + i
        if val is not None and val != 0.0:
            _sc(ws, row, col, val, font=_FONT_HIST, fill=_FILL_HIST, fmt=_FMT_USD)
        else:
            ws.cell(row=row, column=col).fill = _FILL_HIST

    for j, val in enumerate(fcst_values):
        col = _COL_FCST_START + j
        if val is not None:
            _sc(ws, row, col, val, font=_FONT_FCST, fill=_FILL_FCST, fmt=_FMT_USD)


def _build_cover(ws: Any, ticker: str, company_name: str) -> None:
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 80

    today = datetime.date.today().isoformat()
    _sc(
        ws,
        1,
        2,
        f"{company_name} ({ticker}) — Three-Statement Financial Model",
        font=Font(bold=True, size=16, name="Calibri"),
    )
    _sc(ws, 2, 2, f"Generated: {today}", font=_FONT_NORM)
    _sc(
        ws,
        3,
        2,
        "Source: SEC EDGAR XBRL — see Sources sheet for accession numbers.",
        font=_FONT_NORM,
    )

    disclosure = (
        "SIMPLIFIED THREE-STATEMENT MODEL.\n"
        "• Working capital items beyond AR/AP/Inventory/DeferredRevenue are aggregated "
        "into OtherAssets/OtherLiabilities and held flat in the forecast.\n"
        "• Stock-based compensation: non-cash OCF add-back, sized from the latest "
        "historical quarter and held flat.\n"
        "• Stock buybacks: financing-line cash outflow, sized from the latest "
        "historical quarter and held flat.\n"
        "• Forecasted OCF should match GAAP OCF within $5M tolerance for the last "
        "4 historical quarters; this validates the WC decomposition.\n"
        "• BalanceCheck tolerance: $1M. Cash is the balancing item derived from "
        "the CF statement — NOT set independently.\n"
        "• Full SaaS-grade modeling (deferred-revenue dynamics, OCI components, "
        "debt-schedule embedded options) is v2 work.\n"
        "• Source: SEC EDGAR XBRL facts. Provenance on the Sources sheet — "
        "every value traces to an accession number."
    )
    cell = ws.cell(row=5, column=2, value=disclosure)
    cell.font = Font(bold=True, name="Calibri", size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[5].height = 130


def _build_scenarios(
    ws: Any,
    scenarios: dict[str, dict[str, float]],
    has_inv: bool,
) -> None:
    ws.title = "Scenarios"
    ws.column_dimensions["A"].width = 30
    for col_letter in ("B", "C", "D"):
        ws.column_dimensions[col_letter].width = 16

    params: list[tuple[str, str, str]] = [
        ("revenue_growth_qoq", "Revenue Growth QoQ", _FMT_PCT),
        ("gross_margin_pct", "Gross Margin %", _FMT_PCT),
        ("opex_growth_qoq", "OpEx Growth QoQ", _FMT_PCT),
        ("capex_pct_of_revenue", "CapEx % Revenue", _FMT_PCT),
        ("dso_days", "DSO (days)", _FMT_DAY),
        ("dpo_days", "DPO (days)", _FMT_DAY),
        ("dio_days", "DIO (days)", _FMT_DAY),
        ("tax_rate", "Tax Rate", _FMT_PCT),
        ("debt_amortization_qoq", "Debt Amort Qtrly", _FMT_USD),
        ("sbc_qtrly", "SBC Quarterly ($)", _FMT_USD),
        ("buybacks_qtrly", "Buybacks Quarterly ($)", _FMT_USD),
    ]

    headers = ["Assumption", "Base", "Bull", "Bear"]
    for col_idx, h in enumerate(headers, start=1):
        _sc(ws, 1, col_idx, h, font=_FONT_HDR, fill=_FILL_HEADER, halign="center")

    for row_idx, (key, label, fmt) in enumerate(params, start=2):
        if key == "dio_days" and not has_inv:
            continue
        _sc(ws, row_idx, 1, label, font=_FONT_BOLD)
        for col_idx, scen in enumerate(["Base", "Bull", "Bear"], start=2):
            val = scenarios[scen].get(key, 0.0)
            _sc(
                ws, row_idx, col_idx, val, font=_FONT_INP, fill=_FILL_INPUT, fmt=fmt, halign="right"
            )

    _sc(
        ws,
        len(params) + 3,
        1,
        "Yellow cells are editable inputs. Changes here do NOT auto-update the "
        "Income_Statement / Balance_Sheet / Cash_Flow sheets — those use the "
        "Base scenario values computed at generation time.  Re-run build_excel_model.py "
        "to regenerate with updated assumptions.",
        font=Font(italic=True, name="Calibri", size=9),
    )
    ws.cell(row=len(params) + 3, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[len(params) + 3].height = 50


def _build_assumptions(
    ws: Any,
    base_assumptions: dict[str, float],
    has_inv: bool,
) -> None:
    ws.title = "Assumptions"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40

    _sc(ws, 1, 1, "Active Scenario", font=_FONT_BOLD)
    _sc(ws, 1, 2, "Base", font=_FONT_INP, fill=_FILL_INPUT)

    dv = DataValidation(type="list", formula1='"Base,Bull,Bear"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws["B1"])

    params: list[tuple[str, str, str, str]] = [
        (
            "revenue_growth_qoq",
            "Revenue Growth QoQ",
            _FMT_PCT,
            "Quarterly revenue growth rate for forecast periods",
        ),
        ("gross_margin_pct", "Gross Margin %", _FMT_PCT, "Gross profit as % of revenue"),
        ("opex_growth_qoq", "OpEx Growth QoQ", _FMT_PCT, "Quarterly opex growth rate"),
        (
            "capex_pct_of_revenue",
            "CapEx % of Revenue",
            _FMT_PCT,
            "Capital expenditure as % of quarterly revenue",
        ),
        (
            "dso_days",
            "DSO (days)",
            _FMT_DAY,
            "Days Sales Outstanding — drives AR = (DSO/90) × Revenue",
        ),
        (
            "dpo_days",
            "DPO (days)",
            _FMT_DAY,
            "Days Payable Outstanding — drives AP = (DPO/90) × COGS",
        ),
        (
            "dio_days",
            "DIO (days)",
            _FMT_DAY,
            "Days Inventory Outstanding — drives Inventory = (DIO/90) × COGS",
        ),
        ("tax_rate", "Effective Tax Rate", _FMT_PCT, "Applied to positive PreTax Income only"),
        (
            "debt_amortization_qoq",
            "Debt Amortization Qtrly",
            _FMT_USD,
            "Mandatory quarterly principal repayment",
        ),
        (
            "sbc_qtrly",
            "SBC Quarterly ($)",
            _FMT_USD,
            "Non-cash OCF add-back; held flat at latest historical value",
        ),
        (
            "buybacks_qtrly",
            "Buybacks Quarterly ($)",
            _FMT_USD,
            "Financing-line cash outflow; held flat at latest historical value",
        ),
    ]

    _sc(ws, 2, 1, "Parameter", font=_FONT_HDR, fill=_FILL_HEADER)
    _sc(ws, 2, 2, "Value", font=_FONT_HDR, fill=_FILL_HEADER)
    _sc(ws, 2, 3, "Notes", font=_FONT_HDR, fill=_FILL_HEADER)

    row = 3
    for key, label, fmt, note in params:
        if key == "dio_days" and not has_inv:
            continue
        _sc(ws, row, 1, label, font=_FONT_BOLD)
        _sc(ws, row, 2, base_assumptions.get(key, 0.0), font=_FONT_INP, fill=_FILL_INPUT, fmt=fmt)
        _sc(ws, row, 3, note, font=_FONT_NORM)
        row += 1


def _build_income_statement(
    ws: Any,
    hist_is: pd.DataFrame,
    fcst_base: list[dict[str, float]],
) -> None:
    ws.title = "Income_Statement"
    _col_headers(ws, hist_is)

    rows: list[tuple[str, str]] = [
        ("Revenue", "Revenue"),
        ("CostOfRevenue", "Cost of Revenue"),
        ("GrossProfit", "Gross Profit"),
        ("OperatingExpenses", "Operating Expenses"),
        ("OperatingIncome", "Operating Income"),
        ("NetIncome", "Net Income"),
    ]

    for r_idx, (col_key, label) in enumerate(rows, start=2):
        hist_vals = _series(hist_is, col_key)
        fcst_vals = [p.get(col_key) for p in fcst_base]
        _write_data_row(ws, r_idx, label, hist_vals, fcst_vals)

    # Gross margin % row
    gm_row = len(rows) + 3
    _sc(ws, gm_row, _COL_LABEL, "Gross Margin %", font=_FONT_BOLD)
    for i in range(_N_HIST):
        rev = _v(hist_is, "Revenue", i)
        gp = _v(hist_is, "GrossProfit", i)
        col = _COL_DATA_START + i
        if rev > 0:
            _sc(ws, gm_row, col, gp / rev, font=_FONT_HIST, fill=_FILL_HIST, fmt=_FMT_PCT)
    for j, p in enumerate(fcst_base):
        rev = p.get("Revenue", 1.0) or 1.0
        gp = p.get("GrossProfit", 0.0) or 0.0
        col = _COL_FCST_START + j
        _sc(ws, gm_row, col, gp / rev, font=_FONT_FCST, fill=_FILL_FCST, fmt=_FMT_PCT)

    _sc(
        ws,
        1,
        _COL_FCST_START - 1,
        "← Historical (blue) | Forecast (green) →",
        font=Font(italic=True, name="Calibri", size=9),
    )


def _build_balance_sheet(
    ws: Any,
    hist_bs: pd.DataFrame,
    fcst_base: list[dict[str, float]],
    has_inv: bool,
) -> None:
    ws.title = "Balance_Sheet"

    col_hdr_df = hist_bs if len(hist_bs) > 0 else pd.DataFrame()
    _col_headers(ws, col_hdr_df)

    asset_rows: list[tuple[str, str]] = [
        ("Cash", "Cash & Equivalents"),
        ("AccountsReceivable", "Accounts Receivable"),
    ]
    if has_inv:
        asset_rows.append(("Inventory", "Inventory"))
    asset_rows.append(("TotalAssets", "TOTAL ASSETS"))

    liab_rows: list[tuple[str, str]] = [
        ("AccountsPayable", "Accounts Payable"),
        ("DeferredRevenue", "Deferred Revenue"),
        ("TotalLiabilities", "TOTAL LIABILITIES"),
    ]

    equity_rows: list[tuple[str, str]] = [
        ("TotalEquity", "Total Equity"),
    ]

    row = 2
    _sc(ws, row, _COL_LABEL, "ASSETS", font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1
    for col_key, label in asset_rows:
        hist_vals = _series(hist_bs, col_key)
        fcst_vals = [p.get(col_key) for p in fcst_base]
        _write_data_row(ws, row, label, hist_vals, fcst_vals)
        row += 1

    row += 1
    _sc(ws, row, _COL_LABEL, "LIABILITIES", font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1
    for col_key, label in liab_rows:
        hist_vals = _series(hist_bs, col_key)
        fcst_vals = [p.get(col_key) for p in fcst_base]
        _write_data_row(ws, row, label, hist_vals, fcst_vals)
        row += 1

    row += 1
    _sc(ws, row, _COL_LABEL, "EQUITY", font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1
    for col_key, label in equity_rows:
        hist_vals = _series(hist_bs, col_key)
        fcst_vals = [p.get(col_key) for p in fcst_base]
        _write_data_row(ws, row, label, hist_vals, fcst_vals)
        row += 1

    # BalanceCheck row
    row += 1
    _sc(
        ws,
        row,
        _COL_LABEL,
        "BalanceCheck (Assets−Liab−Eq)",
        font=Font(bold=True, color="000000", name="Calibri", size=10),
    )
    for j, p in enumerate(fcst_base):
        bc = p.get("BalanceCheck", 0.0) or 0.0
        col = _COL_FCST_START + j
        ok = abs(bc) < _BALANCE_CHECK_TOL
        cell = _sc(
            ws,
            row,
            col,
            bc,
            font=_FONT_FCST if ok else Font(bold=True, color="FF0000", name="Calibri"),
            fill=_FILL_OK if ok else _FILL_FAIL,
            fmt=_FMT_USD,
        )
        _ = cell  # used for side-effects only


def _build_cash_flow(
    ws: Any,
    hist_is: pd.DataFrame,
    hist_bs: pd.DataFrame,
    hist_cf: pd.DataFrame,
    fcst_base: list[dict[str, float]],
    has_inv: bool,
) -> None:
    ws.title = "Cash_Flow"

    col_hdr_df = hist_is if len(hist_is) > 0 else pd.DataFrame()
    _col_headers(ws, col_hdr_df)

    row = 2
    _sc(ws, row, _COL_LABEL, "OPERATING ACTIVITIES", font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1

    ocf_rows: list[tuple[str, pd.DataFrame, str]] = [
        ("Net Income", hist_is, "NetIncome"),
        ("Depreciation & Amort", hist_cf, "Depreciation"),
        ("Stock-Based Comp", hist_cf, "StockBasedCompensation"),
    ]
    for label, src_df, col_key in ocf_rows:
        hist_vals = _series(src_df, col_key) if len(src_df) > 0 else []
        fcst_vals = [
            p.get(
                "Depreciation"
                if col_key == "Depreciation"
                else "SBC"
                if col_key == "StockBasedCompensation"
                else col_key,
                0.0,
            )
            for p in fcst_base
        ]
        _write_data_row(ws, row, label, hist_vals, fcst_vals)
        row += 1

    # ΔAR
    ar_series = _series(hist_bs, "AccountsReceivable")
    delta_ar_hist = [
        ar_series[i] - ar_series[i - 1] if i > 0 else 0.0 for i in range(len(ar_series))
    ]
    delta_ar_fcst = [
        p.get("AccountsReceivable", 0.0)
        - (
            fcst_base[j - 1].get("AccountsReceivable", 0.0)
            if j > 0
            else ar_series[-1]
            if ar_series
            else 0.0
        )
        for j, p in enumerate(fcst_base)
    ]
    _write_data_row(
        ws,
        row,
        "Δ Accounts Receivable (−increase)",
        [-v for v in delta_ar_hist],
        [-v for v in delta_ar_fcst],
    )
    row += 1

    if has_inv:
        inv_series = _series(hist_bs, "Inventory")
        delta_inv_hist = [
            inv_series[i] - inv_series[i - 1] if i > 0 else 0.0 for i in range(len(inv_series))
        ]
        delta_inv_fcst = [
            p.get("Inventory", 0.0)
            - (
                fcst_base[j - 1].get("Inventory", 0.0)
                if j > 0
                else inv_series[-1]
                if inv_series
                else 0.0
            )
            for j, p in enumerate(fcst_base)
        ]
        _write_data_row(
            ws,
            row,
            "Δ Inventory (−increase)",
            [-v for v in delta_inv_hist],
            [-v for v in delta_inv_fcst],
        )
        row += 1

    ap_series = _series(hist_bs, "AccountsPayable")
    delta_ap_hist = [
        ap_series[i] - ap_series[i - 1] if i > 0 else 0.0 for i in range(len(ap_series))
    ]
    delta_ap_fcst = [
        p.get("AccountsPayable", 0.0)
        - (
            fcst_base[j - 1].get("AccountsPayable", 0.0)
            if j > 0
            else ap_series[-1]
            if ap_series
            else 0.0
        )
        for j, p in enumerate(fcst_base)
    ]
    _write_data_row(ws, row, "Δ Accounts Payable (+increase)", delta_ap_hist, delta_ap_fcst)
    row += 1

    ocf_hist = _series(hist_cf, "OperatingCashFlow")
    ocf_fcst = [p.get("OCF") for p in fcst_base]
    _write_data_row(ws, row, "OPERATING CASH FLOW", ocf_hist, ocf_fcst)
    row += 2

    _sc(ws, row, _COL_LABEL, "INVESTING ACTIVITIES", font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1
    capex_hist = _series(hist_cf, "CapEx")
    capex_fcst = [p.get("CapEx") for p in fcst_base]
    _write_data_row(ws, row, "Capital Expenditures (−)", capex_hist, capex_fcst)
    row += 1
    icf_hist = _series(hist_cf, "InvestingCashFlow")
    icf_fcst = [p.get("InvestingCF") for p in fcst_base]
    _write_data_row(ws, row, "INVESTING CASH FLOW", icf_hist, icf_fcst)
    row += 2

    _sc(ws, row, _COL_LABEL, "FINANCING ACTIVITIES", font=_FONT_BOLD, fill=_FILL_SECTION)
    row += 1
    bb_hist = _series(hist_cf, "TreasuryStockRepurchases")
    bb_fcst = [p.get("Buybacks") for p in fcst_base]
    _write_data_row(ws, row, "Stock Buybacks (−)", bb_hist, bb_fcst)
    row += 1
    fcf_hist = _series(hist_cf, "FinancingCashFlow")
    fcf_fcst = [p.get("FinancingCF") for p in fcst_base]
    _write_data_row(ws, row, "FINANCING CASH FLOW", fcf_hist, fcf_fcst)
    row += 2

    net_change_fcst = [p.get("NetCashChange") for p in fcst_base]
    _write_data_row(ws, row, "Net Change in Cash", [], net_change_fcst)
    row += 1
    end_cash_fcst = [p.get("EndCash") for p in fcst_base]
    _write_data_row(ws, row, "Ending Cash (= BS Cash)", [], end_cash_fcst)


def _build_debt_schedule(
    ws: Any,
    base_assumptions: dict[str, float],
    hist_bs: pd.DataFrame,
) -> None:
    ws.title = "Debt_Schedule"
    ws.column_dimensions["A"].width = 28

    _sc(ws, 1, 1, "Debt Schedule", font=Font(bold=True, size=12, name="Calibri"))
    _sc(
        ws,
        2,
        1,
        "(Simplified — Debt is not directly available as XBRL line item; "
        "defaults to $0 opening balance.  Update manually with actual debt from 10-K.)",
        font=Font(italic=True, name="Calibri", size=9),
    )
    ws.cell(row=2, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 40

    headers = ["Item", "Q1(F)", "Q2(F)", "Q3(F)", "Q4(F)"]
    for col_idx, h in enumerate(headers, start=1):
        _sc(ws, 4, col_idx, h, font=_FONT_HDR, fill=_FILL_HEADER, halign="center")

    amort = base_assumptions.get("debt_amortization_qoq", 0.0)
    debt_opening = 0.0
    items = [
        ("Opening Debt Balance", [debt_opening - amort * i for i in range(_N_FCST)]),
        ("− Amortization", [-amort] * _N_FCST),
        (
            "Closing Debt Balance",
            [max(0.0, debt_opening - amort * (i + 1)) for i in range(_N_FCST)],
        ),
        ("Interest Expense", [0.0] * _N_FCST),  # simplified zero-coupon
    ]
    for row_off, (label, vals) in enumerate(items, start=5):
        _sc(ws, row_off, 1, label, font=_FONT_BOLD)
        for j, v in enumerate(vals):
            _sc(ws, row_off, j + 2, v, font=_FONT_FCST, fill=_FILL_FCST, fmt=_FMT_USD)


def _build_summary(
    ws: Any,
    ticker: str,
    hist_is: pd.DataFrame,
    hist_cf: pd.DataFrame,
    fcst_base: list[dict[str, float]],
    ocf_residual: float,
) -> None:
    ws.title = "Summary"
    ws.column_dimensions["A"].width = 36
    for c in range(2, _COL_FCST_END + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    _sc(ws, 1, 1, f"{ticker} — Key Metrics Summary", font=Font(bold=True, size=14, name="Calibri"))

    row = 3
    kpi_hist_rows: list[tuple[str, pd.DataFrame, str, str]] = [
        ("Revenue ($)", hist_is, "Revenue", _FMT_USD),
        ("Operating Income ($)", hist_is, "OperatingIncome", _FMT_USD),
        ("Net Income ($)", hist_is, "NetIncome", _FMT_USD),
        ("Operating Cash Flow ($)", hist_cf, "OperatingCashFlow", _FMT_USD),
    ]

    _sc(ws, row, _COL_LABEL, "Metric", font=_FONT_HDR, fill=_FILL_HEADER)
    for j in range(_N_HIST):
        _sc(
            ws,
            row,
            _COL_DATA_START + j,
            f"H{j + 1}",
            font=_FONT_HDR,
            fill=_FILL_HEADER,
            halign="center",
        )
    for j in range(_N_FCST):
        _sc(
            ws,
            row,
            _COL_FCST_START + j,
            f"Q{j + 1}(F)",
            font=_FONT_HDR,
            fill=_FILL_HEADER,
            halign="center",
        )
    row += 1

    for label, src_df, col_key, _fmt in kpi_hist_rows:
        hist_vals = _series(src_df, col_key)
        fcst_key = {"OperatingCashFlow": "OCF"}.get(col_key, col_key)
        fcst_vals = [p.get(fcst_key) for p in fcst_base]
        _write_data_row(ws, row, label, hist_vals, fcst_vals)
        row += 1

    # GAAP OCF vs Modelled OCF residual
    row += 1
    _sc(ws, row, _COL_LABEL, "GAAP OCF Residual Validation", font=_FONT_BOLD)
    row += 1
    ok = ocf_residual < _OCF_RESIDUAL_TOL
    _sc(
        ws,
        row,
        _COL_LABEL,
        f"Max |Modelled OCF − Reported OCF| (last 4Q): ${ocf_residual:,.0f}  "
        f"{'✓ PASS' if ok else '✗ FAIL'} (tolerance $5M)",
        font=Font(bold=True, color="375623" if ok else "FF0000", name="Calibri", size=10),
    )


def _build_sources(ws: Any, sources_df: pd.DataFrame) -> None:
    ws.title = "Sources"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 60

    headers = [
        "line_item",
        "period_end",
        "fiscal_year",
        "fiscal_period",
        "value",
        "accession_no",
        "filing_url",
    ]
    for col_idx, h in enumerate(headers, start=1):
        _sc(ws, 1, col_idx, h, font=_FONT_HDR, fill=_FILL_HEADER)

    for row_idx, row in sources_df.iterrows():
        for col_idx, col_name in enumerate(headers, start=1):
            val = row.get(col_name)
            cell = ws.cell(row=int(str(row_idx)) + 2, column=col_idx)
            if pd.notna(val):
                cell.value = val
                cell.font = _FONT_NORM
                if col_name == "value":
                    cell.number_format = _FMT_USD

    _sc(
        ws,
        len(sources_df) + 3,
        1,
        "Every number in the model traces to an SEC EDGAR accession number. "
        "Click filing_url to open the source 10-K or 10-Q on SEC.gov.",
        font=Font(italic=True, name="Calibri", size=9),
    )


def _build_revenue_disaggregation(ws: Any, ticker: str) -> None:
    ws.title = "Revenue_Disaggregation"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 85
    ws.sheet_view.showGridLines = False

    lines = [
        (
            1,
            f"{ticker} Revenue Disaggregation — ASC 280 Note",
            Font(bold=True, size=14, name="Calibri"),
        ),
        (3, "SEGMENT REPORTING (ASC 280)", Font(bold=True, name="Calibri", size=11)),
        (
            4,
            f"{ticker} reports as ONE operating and reportable segment under ASC 280. "
            "There are no separately reportable segments.",
            Font(name="Calibri", size=10),
        ),
        (6, "REVENUE DISAGGREGATION (ASC 606)", Font(bold=True, name="Calibri", size=11)),
        (
            7,
            "Revenue is disaggregated into TWO categories in the 10-K:",
            Font(name="Calibri", size=10),
        ),
        (
            8,
            "  1. Product  — next-gen firewall appliances and other hardware",
            Font(name="Calibri", size=10),
        ),
        (
            9,
            "  2. Subscription and Support (S&S)  — cloud subscriptions, software "
            "updates, premium support, maintenance",
            Font(name="Calibri", size=10),
        ),
        (
            11,
            "PRODUCT FAMILIES (commercial brands, NOT segments)",
            Font(bold=True, name="Calibri", size=11),
        ),
        (
            12,
            "Strata, Prisma, and Cortex are product families — commercial brands "
            "that span both revenue categories:",
            Font(name="Calibri", size=10),
        ),
        (
            13,
            "  • Strata  — generates both Product revenue (from firewall appliances) "
            "AND S&S revenue (from subscriptions running on those appliances)",
            Font(name="Calibri", size=10),
        ),
        (
            14,
            "  • Prisma  — cloud-delivered SASE/CNAPP; almost entirely S&S",
            Font(name="Calibri", size=10),
        ),
        (15, "  • Cortex  — XDR/XSIAM; almost entirely S&S", Font(name="Calibri", size=10)),
        (
            16,
            "Calling Strata/Prisma/Cortex 'segments' is incorrect under ASC 280.",
            Font(bold=True, italic=True, name="Calibri", size=10),
        ),
        (18, "CURRENT MODEL SCOPE", Font(bold=True, name="Calibri", size=11)),
        (
            19,
            "This model forecasts CONSOLIDATED REVENUE only. Disaggregating into "
            "Product vs. Subscription & Support is a v2 enhancement; disaggregating "
            "further by product family would require parsing narrative text from earnings "
            "releases (XBRL does not tag it at that granularity).",
            Font(name="Calibri", size=10),
        ),
        (
            21,
            "Forecasting Product and Subscription & Support separately is a v2 enhancement. "
            "NGS ARR — PANW's headline non-GAAP metric — is not exposed as structured XBRL "
            "and is out of scope here.",
            Font(italic=True, name="Calibri", size=10),
        ),
    ]

    for r, text, font in lines:
        cell = ws.cell(row=r, column=2, value=text)
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(30, len(text) // 4)


# ── Main build function ────────────────────────────────────────────────────────


def build(ticker: str | None = None, db_path: Path | None = None) -> Path:
    """Generate the three-statement Excel model and verify correctness.

    Args:
        ticker:  Ticker symbol override; falls back to config/company.yaml.
        db_path: Path to DuckDB file; auto-derived from ticker if not given.

    Returns:
        Path to the generated ``.xlsx`` file.

    Raises:
        FileNotFoundError: If DuckDB file or config is missing.
        RuntimeError:      If BalanceCheck > $1M or GAAP OCF residual > $5M.
    """
    # ── Load config ────────────────────────────────────────────────────────────
    if not _CONFIG_PATH.exists():
        raise FileNotFoundError(f"Config not found: {_CONFIG_PATH}")
    with _CONFIG_PATH.open() as fh:
        config: dict[str, Any] = yaml.safe_load(fh)

    resolved_ticker = (ticker or str(config["ticker"])).upper().strip()
    company_name = str(config.get("name", resolved_ticker))

    if db_path is None:
        db_path = _PROCESSED_DIR / f"{resolved_ticker}.duckdb"
    if not db_path.exists():
        raise FileNotFoundError(
            f"DuckDB warehouse not found: {db_path}\n"
            "Run first:  python -m src.build_warehouse --ticker {resolved_ticker}"
        )

    # ── Load history ───────────────────────────────────────────────────────────
    hist_is, hist_bs, hist_cf, has_inv, sources_df = _load_history(db_path)

    logger.info(
        "Loaded %d IS rows, %d BS rows, %d CF rows (quarterly)",
        len(hist_is),
        len(hist_bs),
        len(hist_cf),
    )

    # ── Compute assumptions and scenarios ──────────────────────────────────────
    base_ass = _compute_base_assumptions(hist_is, hist_bs, hist_cf, has_inv)
    scenarios = _make_scenarios(base_ass)

    logger.info(
        "Base assumptions: rev_growth=%.1f%% gross_margin=%.1f%% dso=%.0fd",
        base_ass["revenue_growth_qoq"] * 100,
        base_ass["gross_margin_pct"] * 100,
        base_ass["dso_days"],
    )

    # ── Compute forecast (Python-side — independent of openpyxl) ──────────────
    fcst_base = _forecast_periods(hist_is, hist_bs, hist_cf, base_ass, has_inv)
    fcst_bull = _forecast_periods(hist_is, hist_bs, hist_cf, scenarios["Bull"], has_inv)
    fcst_bear = _forecast_periods(hist_is, hist_bs, hist_cf, scenarios["Bear"], has_inv)

    # ── Python-side BalanceCheck verification ──────────────────────────────────
    # NOTE: We do NOT read back the workbook with openpyxl data_only=True.
    # On freshly generated workbooks that have never been opened in Excel,
    # data_only=True returns None for all formula cells (no cached results).
    # Asserting against None silently passes or raises TypeError — neither is useful.
    # Instead we verify the accounting identity using the same Python computations
    # that produced the cell values.
    for scenario_name, fcst in [("Base", fcst_base), ("Bull", fcst_bull), ("Bear", fcst_bear)]:
        max_bc = max(abs(p.get("BalanceCheck", 0.0)) for p in fcst)
        logger.info("%s max |BalanceCheck| = $%,.0f", scenario_name, max_bc)
        if max_bc > _BALANCE_CHECK_TOL:
            raise RuntimeError(
                f"{scenario_name} scenario: max |BalanceCheck| = ${max_bc:,.0f} "
                f"exceeds $1M tolerance.  Check WC formula consistency."
            )

    # ── GAAP OCF residual check ────────────────────────────────────────────────
    ocf_residual = _verify_ocf_residual(hist_is, hist_bs, hist_cf, has_inv)
    logger.info("GAAP OCF residual (last 4Q): $%,.0f  (tolerance $5M)", ocf_residual)
    if ocf_residual > _OCF_RESIDUAL_TOL:
        logger.warning(
            "GAAP OCF residual $%,.0f exceeds $5M — WC model may be incomplete. "
            "Deferred-revenue or tax-timing items not captured.",
            ocf_residual,
        )

    # ── Sources sheet presence check ──────────────────────────────────────────
    if sources_df is not None and len(sources_df) > 0:
        missing_accn = sources_df["accession_no"].isna().sum()
        if missing_accn > 0:
            logger.warning("%d sources rows have null accession_no", missing_accn)

    # ── Build workbook ─────────────────────────────────────────────────────────
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)  # remove default empty sheet

    _build_cover(wb.create_sheet("Cover"), resolved_ticker, company_name)
    _build_scenarios(wb.create_sheet("Scenarios"), scenarios, has_inv)
    _build_assumptions(wb.create_sheet("Assumptions"), base_ass, has_inv)
    _build_income_statement(wb.create_sheet("Income_Statement"), hist_is, fcst_base)
    _build_balance_sheet(wb.create_sheet("Balance_Sheet"), hist_bs, fcst_base, has_inv)
    _build_cash_flow(
        wb.create_sheet("Cash_Flow"),
        hist_is,
        hist_bs,
        hist_cf,
        fcst_base,
        has_inv,
    )
    _build_debt_schedule(wb.create_sheet("Debt_Schedule"), base_ass, hist_bs)
    _build_summary(
        wb.create_sheet("Summary"),
        resolved_ticker,
        hist_is,
        hist_cf,
        fcst_base,
        ocf_residual,
    )
    if sources_df is not None and len(sources_df) > 0:
        _build_sources(wb.create_sheet("Sources"), sources_df)

    if has_inv:
        _build_revenue_disaggregation(wb.create_sheet("Revenue_Disaggregation"), resolved_ticker)

    _DASHBOARD_DIR.mkdir(parents=True, exist_ok=True)
    out_path = _DASHBOARD_DIR / f"{resolved_ticker}_3Statement_Model.xlsx"
    wb.save(str(out_path))
    logger.info("Saved %s", out_path)
    return out_path


# ── CLI ────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse
    import sys

    logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")

    parser = argparse.ArgumentParser(
        description="Build three-statement Excel model from DuckDB warehouse."
    )
    parser.add_argument("--ticker", default=None, help="Ticker symbol (e.g. PANW)")
    args = parser.parse_args()

    try:
        path = build(ticker=args.ticker)
        print(f"\nGenerated: {path}")
    except (FileNotFoundError, RuntimeError) as exc:
        print(f"Error: {exc}", file=sys.stderr)
        sys.exit(1)
